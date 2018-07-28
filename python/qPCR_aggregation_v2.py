import warnings
import os
import fnmatch
import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from pathlib import Path
import json
import sys
import time

new_wb = Workbook()
retest_list = []


# Modified date format, rounding and other
def qpcr_qc_summary(folder_name):
    file_list = qpcr_extract_excels(folder_name)

    # create a new workbook , and initialize its worksheet name, header
    new_ws_qc = new_wb.create_sheet(title="QCs Summary", index=0)
    new_ws_qc.cell(row=1, column=1).value = "PCRRunNumber"
    new_ws_qc.cell(row=1, column=2).value = "ExtractionDate"
    new_ws_qc.cell(row=1, column=3).value = "SampleName"
    new_ws_qc.cell(row=1, column=4).value = "WellPosition"
    new_ws_qc.cell(row=1, column=5).value = "CtMean"
    new_ws_qc.cell(row=1, column=6).value = "CtSD"
    new_ws_qc.cell(row=1, column=7).value = "QuantityMeanPer10uL"
    new_ws_qc.cell(row=1, column=8).value = "QuantitySDPer10uL"
    new_ws_qc.cell(row=1, column=9).value = "QuantityCVPer10uL"
    new_ws_qc.cell(row=1, column=10).value = "QuantityNominalPer10uL"
    new_ws_qc.cell(row=1, column=11).value = "PercentRE"
    new_ws_qc.cell(row=1, column=12).value = "QC"

    # create a counter for keep track on row number for the new xlsx file
    row_counter = 2

    # iterate through all data spreadsheets and copy 4 values (run number, slope, intercept and R^2)
    # into the new file
    for file_name in file_list:
        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['QCs']
        row_start = 10
        while cur_ws.cell(row=row_start, column=1).value is not None:
            for j in range(1, 13, 1):
                if j == 1:
                    new_ws_qc.cell(row=row_counter, column=j).value = cur_ws['B3'].value
                elif j == 2:
                    extraction_date = str(cur_ws['D2'].value).split()
                    formal_date = datetime.datetime.strptime(extraction_date[0], '%Y-%m-%d').strftime('%d%b%Y')
                    new_ws_qc.cell(row=row_counter, column=j).value = formal_date
                elif j == 5 or j == 6:
                    if cur_ws.cell(row=row_start, column=j - 2).value == "Undetermined":
                        new_ws_qc.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 2).value
                    else:
                        new_ws_qc.cell(row=row_counter, column=j).value = round(
                            float(cur_ws.cell(row=row_start, column=j - 2).value), 3)
                elif j == 7 or j == 8:
                    new_ws_qc.cell(row=row_counter, column=j).value = round(
                        float(cur_ws.cell(row=row_start, column=j - 2).value))
                elif j == 9:
                    if cur_ws.cell(row=row_start, column=j - 2).value is not None:
                        num = round((float(cur_ws.cell(row=row_start, column=j - 2).value) * 100), 2)
                        new_ws_qc.cell(row=row_counter, column=j).value = str(num) + "%"
                    else:
                        new_ws_qc.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 2).value
                elif j == 11:
                    if cur_ws.cell(row=row_start, column=j - 2).value != "N/A":
                        new_ws_qc.cell(row=row_counter, column=j).value = round(
                            float(cur_ws.cell(row=row_start, column=j - 2).value), 1)
                    else:
                        new_ws_qc.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 2).value
                # elif j == 12:
                #     if qc_list[file_list.index(file_name)]:
                #         new_ws_qc.cell(row=row_counter, column=j).value = "Yes"
                else:
                    new_ws_qc.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 2).value

            row_counter = row_counter + 1
            row_start = row_start + 1
    qc_summary_list = []
    keys = []
    for j in range(1, 13, 1):
        keys.append(new_ws_qc.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 13, 1):
            row_data[keys[j - 1]] = new_ws_qc.cell(row=row_number, column=j).value
        qc_summary_list.append(row_data)
    print(json.dumps(qc_summary_list))

    # os.chdir(savepath)
    # timestr = time.strftime("%Y-%m-%d_%H%M%S")
    # new_wb.save('qPCR_QC_Summary.xlsx')
    # new_wb.save(filename + 'qPCR_QC_Summary.xlsx')


def qpcr_raw_data(folder_name, command):
    file_list = qpcr_extract_excels(folder_name)

    new_ws_sr = new_wb.create_sheet(title="Raw Data Aggregation", index=0)
    new_ws_sr.cell(row=1, column=1).value = "ExtractionNumber"
    new_ws_sr.cell(row=1, column=2).value = "PCRRunNumber"
    new_ws_sr.cell(row=1, column=3).value = "ExtractionSampleNumber"
    new_ws_sr.cell(row=1, column=4).value = "PunchNumber"
    new_ws_sr.cell(row=1, column=5).value = "AnimalID"
    new_ws_sr.cell(row=1, column=6).value = "TissueorSampleType"
    new_ws_sr.cell(row=1, column=7).value = "CollectionDate"
    new_ws_sr.cell(row=1, column=8).value = "DNAPerrxn"
    new_ws_sr.cell(row=1, column=9).value = "SampleName"
    new_ws_sr.cell(row=1, column=10).value = "WellPosition"
    new_ws_sr.cell(row=1, column=11).value = "CtMean"
    new_ws_sr.cell(row=1, column=12).value = "CtSD"
    new_ws_sr.cell(row=1, column=13).value = "QuantityMean"
    new_ws_sr.cell(row=1, column=14).value = "QuantitySD"
    new_ws_sr.cell(row=1, column=15).value = "QtyCVPercent"
    new_ws_sr.cell(row=1, column=16).value = "CNPerug"
    new_ws_sr.cell(row=1, column=17).value = "Flag"
    new_ws_sr.cell(row=1, column=18).value = "QC"

    row_counter = 2
    for file_name in file_list:
        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['Sample result']
        row_start = 10
        while cur_ws.cell(row=row_start, column=1).value != "N/A" and cur_ws.cell(row=row_start,
                                                                                  column=1).value is not None:
            retest = False
            for j in range(1, 20, 1):
                # print(row_start, j)
                if j == 7:
                    collection_date = str(cur_ws.cell(row=row_start, column=j).value).split()
                    formal_date = datetime.datetime.strptime(collection_date[0], '%Y-%m-%d').strftime('%d%b%Y')
                    new_ws_sr.cell(row=row_counter, column=j).value = formal_date
                elif j == 11 or j == 12:
                    if cur_ws.cell(row=row_start, column=j).value == "Undetermined":
                        new_ws_sr.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j).value
                    else:
                        new_ws_sr.cell(row=row_counter, column=j).value = round(
                            float(cur_ws.cell(row=row_start, column=j).value), 2)
                elif j == 13 or j == 14:
                    new_ws_sr.cell(row=row_counter, column=j).value = round(
                        float(cur_ws.cell(row=row_start, column=j).value))
                elif j == 15:
                    if cur_ws.cell(row=row_start, column=j).value == "ND":
                        new_ws_sr.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j).value
                    else:
                        new_ws_sr.cell(row=row_counter, column=j).value = round(
                            float(cur_ws.cell(row=row_start, column=j).value), 1)
                elif j == 16:
                    if cur_ws.cell(row=row_start, column=j).value == "Negative":
                        new_ws_sr.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j).value
                    else:
                        new_ws_sr.cell(row=row_counter, column=j).value = round(
                            float(cur_ws.cell(row=row_start, column=j).value))
                elif j == 17:
                    if cur_ws.cell(row=row_start, column=j).value == "Retest":
                        retest = True
                    new_ws_sr.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j).value
                # elif j == 18:
                #     if qc_list[file_list.index(file_name)]:
                #         new_ws_sr.cell(row=row_counter, column=j).value = "Yes"
                else:
                    new_ws_sr.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j).value
            if retest:
                this_row_info = []
                for col in range(1, 20, 1):
                    this_row_info.append(new_ws_sr.cell(row=row_counter, column=col).value)
                retest_list.append(this_row_info)
            row_counter = row_counter + 1
            row_start = row_start + 1

    rawdata_list = []
    keys = []
    for j in range(1, 20, 1):
        keys.append(new_ws_sr.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 20, 1):
            row_data[keys[j - 1]] = new_ws_sr.cell(row=row_number, column=j).value
        rawdata_list.append(row_data)

    if command == 'qPCRraw':
        print(json.dumps(rawdata_list))
        # os.chdir(savepath)
        # new_wb.save(filename + 'qPCR_Sample_Result_Summary.xlsx')


def qpcr_retest_fun():
    new_ws_r = new_wb.create_sheet(title="Retest information", index=0)
    new_ws_r.cell(row=1, column=1).value = "ExtractionNumber"
    new_ws_r.cell(row=1, column=2).value = "PCRRunNumber"
    new_ws_r.cell(row=1, column=3).value = "ExtractionSampleNumber"
    new_ws_r.cell(row=1, column=4).value = "PunchNumber"
    new_ws_r.cell(row=1, column=5).value = "AnimalID"
    new_ws_r.cell(row=1, column=6).value = "TissueorSampleType"
    new_ws_r.cell(row=1, column=7).value = "CollectionDate"
    new_ws_r.cell(row=1, column=8).value = "DNAPerrxn"
    new_ws_r.cell(row=1, column=9).value = "SampleName"
    new_ws_r.cell(row=1, column=10).value = "WellPosition"
    new_ws_r.cell(row=1, column=11).value = "CtMean"
    new_ws_r.cell(row=1, column=12).value = "CtSD"
    new_ws_r.cell(row=1, column=13).value = "QuantityMean"
    new_ws_r.cell(row=1, column=14).value = "QuantitySD"
    new_ws_r.cell(row=1, column=15).value = "QtyCVPercent"
    new_ws_r.cell(row=1, column=16).value = "CNPerug"
    new_ws_r.cell(row=1, column=17).value = "Flag"
    new_ws_r.cell(row=1, column=18).value = "QC"

    row_counter = 2
    for entry in retest_list:
        for col in range(1, 20, 1):
            new_ws_r.cell(row=row_counter, column=col).value = entry[col - 1]
        row_counter = row_counter + 1

    retest_summary_list = []
    keys = []
    for j in range(1, 20, 1):
        keys.append(new_ws_r.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 20, 1):
            row_data[keys[j - 1]] = new_ws_r.cell(row=row_number, column=j).value
        retest_summary_list.append(row_data)
    print(json.dumps(retest_summary_list))
    # os.chdir(savepath)
    # new_wb.save(filename + 'qPCR_Retest_Summary.xlsx')


def qpcr_each_qc(folder_name):
    file_list = qpcr_extract_excels(folder_name)

    new_ws_cic = new_wb.create_sheet(title="Result for each QC well", index=0)
    row_list = [47, 48, 49, 59, 60, 61, 71, 72, 73, 83, 84, 85, 128, 129, 130]
    row_counter = 1
    for file_name in file_list:
        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['QCs']
        pcr_run_number = cur_ws['B3'].value
        extraction_date = str(cur_ws['D2'].value).split()
        formal_date = datetime.datetime.strptime(extraction_date[0], '%Y-%m-%d').strftime('%d%b%Y')
        cur_ws = cur_wb['raw Results']
        if row_counter == 1:
            for j in range(1, 17, 1):
                if j == 1:
                    new_ws_cic.cell(row=row_counter, column=j).value = "PCRRunNumber"
                elif j == 2:
                    new_ws_cic.cell(row=row_counter, column=j).value = "ExtractionDate"
                elif j == 4:
                    new_ws_cic.cell(row=row_counter, column=j).value = "WellPosition"
                elif j == 6:
                    new_ws_cic.cell(row=row_counter, column=j).value = "SampleName"
                elif j == 7:
                    new_ws_cic.cell(row=row_counter, column=j).value = "TargetName"
                elif j == 12:
                    new_ws_cic.cell(row=row_counter, column=j).value = "CtMean"
                elif j == 13:
                    new_ws_cic.cell(row=row_counter, column=j).value = "CtSD"
                elif j == 15:
                    new_ws_cic.cell(row=row_counter, column=j).value = "QuantityMean"
                elif j == 16:
                    new_ws_cic.cell(row=row_counter, column=j).value = "QuantitySD"
                else:
                    new_ws_cic.cell(row=row_counter, column=j).value = cur_ws.cell(row=43, column=j - 2).value
            row_counter = row_counter + 1

        for row_number in row_list:
            if cur_ws.cell(row=row_number, column=1).value is None:
                continue
            for j in range(1, 17, 1):
                if j == 1:
                    new_ws_cic.cell(row=row_counter, column=j).value = pcr_run_number
                elif j == 2:
                    new_ws_cic.cell(row=row_counter, column=j).value = formal_date
                elif 11 <= j <= 16:
                    if cur_ws.cell(row=row_number, column=j - 2).value == "Undetermined":
                        new_ws_cic.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_number,
                                                                                       column=j - 2).value
                    elif cur_ws.cell(row=row_number, column=j - 2).value == "":
                        continue
                    else:
                        new_ws_cic.cell(row=row_counter, column=j).value = round(float(cur_ws.cell(row=row_number,
                                                                                                   column=j - 2).value),
                                                                                 3)
                else:
                    new_ws_cic.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_number, column=j - 2).value
            row_counter = row_counter + 1

    each_qc_list = []
    keys = []
    for j in range(1, 17, 1):
        keys.append(new_ws_cic.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 17, 1):
            row_data[keys[j - 1]] = new_ws_cic.cell(row=row_number, column=j).value
        each_qc_list.append(row_data)
    print(json.dumps(each_qc_list))
    # os.chdir(savepath)
    # new_wb.save(filename + 'qPCR_QC_Detail_Summary.xlsx')


def qpcr_extract_excels(folder_name):
    file_list = []
    # assign p as the correct directory
    p = Path(folder_name)
    # extract excel files with right leading part directly from directories
    # qc_list = []
    qPCR_file_list = list(p.glob('2377_009*.xlsx'))

    # for qPCR_file in qPCR_file_list:
    #     if fnmatch.fnmatch(str(qPCR_file), '*QC.xlsx'):
    #         qc_list.append(True)
    #     else:
    #         file_list.append(str(qPCR_file))
    # if len(file_list) != len(qc_list):
    #     qc_list.append(False)

    # replace/ignore file if a same file named ***repeated.xlsx
    for qPCR_file in qPCR_file_list:
        if fnmatch.fnmatch(str(qPCR_file), '*repeat.xlsx'):
            batch_name = str(qPCR_file).split()[0] + '.xlsx'
            if batch_name in file_list:
                file_list[file_list.index(batch_name)] = str(qPCR_file)
            if str(qPCR_file) not in file_list:
                file_list.append(str(qPCR_file))
        # skip qc files first
        elif fnmatch.fnmatch(str(qPCR_file), '*QC.xlsx'):
            continue
        else:
            contain_the_file = False
            for single_file in file_list:
                if str(qPCR_file).split('.')[0] == single_file.split()[0]:
                    contain_the_file = True
            if not contain_the_file:
                file_list.append(str(qPCR_file))
    return file_list


def wes_extract_excels(folder_name):
    file_list = []
    # assign p as the correct directory
    p = Path(folder_name)

    CSF_Sample_file_list = list(p.glob('CSF Sample*.xlsx'))
    for CSF_Sample_file in CSF_Sample_file_list:
        file_list.append(str(CSF_Sample_file))

    return file_list


def wes_linear_regression(folder_name):
    file_list = wes_extract_excels(folder_name)

    new_ws_l = new_wb.create_sheet(title="Linear Regression Data Summary", index=0)
    new_ws_l.cell(row=1, column=1).value = "RunNumber"
    new_ws_l.cell(row=1, column=2).value = "Slope"
    new_ws_l.cell(row=1, column=3).value = "Intercept"
    new_ws_l.cell(row=1, column=4).value = "RSquare"

    row_counter = 2
    for file_name in file_list:
        # extract run number
        file_name_slice = file_name.split(".xlsx")
        this_run_number = int(file_name_slice[0].split()[-1])
        if file_name.split()[-3] == "masked":
            this_run_number = str(this_run_number) + " masked"

        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['Data Analysis']
        for j in range(1, 5, 1):
            if j == 1:
                new_ws_l.cell(row=row_counter, column=j).value = this_run_number
            else:
                new_ws_l.cell(row=row_counter, column=j).value = cur_ws.cell(row=7, column=j + 5).value
        row_counter = row_counter + 1

    linear_regression_list = []
    keys = []
    for j in range(1, 5, 1):
        keys.append(new_ws_l.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 5, 1):
            row_data[keys[j - 1]] = new_ws_l.cell(row=row_number, column=j).value
        linear_regression_list.append(row_data)
    print(json.dumps(linear_regression_list))


def wes_standard_curve(folder_name):
    file_list = wes_extract_excels(folder_name)

    new_ws_sc = new_wb.create_sheet(title="Standard Curve Data Summary", index=0)
    new_ws_sc.cell(row=1, column=1).value = "RunNumber"
    new_ws_sc.cell(row=1, column=2).value = "Std"
    new_ws_sc.cell(row=1, column=3).value = "TPP1ConcngPermL"
    new_ws_sc.cell(row=1, column=4).value = "Area"
    new_ws_sc.cell(row=1, column=5).value = "BackCalculatedConcngPermL"
    new_ws_sc.cell(row=1, column=6).value = "PercentRE"

    row_counter = 2
    for file_name in file_list:
        # extract run number
        file_name_slice = file_name.split(".xlsx")
        this_run_number = int(file_name_slice[0].split()[-1])
        if file_name.split()[-3] == "masked":
            this_run_number = str(this_run_number) + " masked"

        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['Data Analysis']
        row_start = 6
        while cur_ws.cell(row=row_start, column=1).value is not None:
            for j in range(1, 7, 1):
                if j == 1:
                    new_ws_sc.cell(row=row_counter, column=j).value = this_run_number
                elif j == 2:
                    new_ws_sc.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 1).value
                else:
                    if cur_ws.cell(row=row_start, column=j - 1).value == "mask" or cur_ws.cell(row=row_start,
                                                                                               column=j - 1).value == "masked":
                        new_ws_sc.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 1).value
                    else:
                        new_ws_sc.cell(row=row_counter, column=j).value = round(
                            float(cur_ws.cell(row=row_start, column=j - 1).value), 3)
            row_start = row_start + 1
            row_counter = row_counter + 1

    standard_curve_list = []
    keys = []
    for j in range(1, 7, 1):
        keys.append(new_ws_sc.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 7, 1):
            row_data[keys[j - 1]] = new_ws_sc.cell(row=row_number, column=j).value
        standard_curve_list.append(row_data)
    print(json.dumps(standard_curve_list))


def wes_upper_and_lower_bond(folder_name):
    file_list = wes_extract_excels(folder_name)

    new_ws_ul = new_wb.create_sheet(title="Upper and Lower bond Summary", index=0)
    new_ws_ul.cell(row=1, column=1).value = "RunNumber"
    new_ws_ul.cell(row=1, column=2).value = "ULOQ"
    new_ws_ul.cell(row=1, column=3).value = "LLOQ"

    row_counter = 2
    for file_name in file_list:
        # extract run number
        file_name_slice = file_name.split(".xlsx")
        this_run_number = int(file_name_slice[0].split()[-1])
        if file_name.split()[-3] == "masked":
            this_run_number = str(this_run_number) + " masked"
        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['Data Analysis']
        row_start = 6
        row_end = 12
        for j in range(1, 4, 1):
            if j == 1:
                new_ws_ul.cell(row=row_counter, column=j).value = this_run_number
            elif j == 2:
                if cur_ws.cell(row=row_start, column=5).value != "masked" and cur_ws.cell(row=row_start,
                                                                                          column=5).value != "mask" and \
                        abs(float(cur_ws.cell(row=row_start, column=5).value)) < 30:
                    new_ws_ul.cell(row=row_counter, column=j).value = round(
                        float(cur_ws.cell(row=row_start, column=j).value - 5), 3)
                elif cur_ws.cell(row=row_start + 1, column=5).value != "masked" and cur_ws.cell(row=row_start + 1,
                                                                                                column=5).value != "mask" \
                        and abs(float(cur_ws.cell(row=row_start + 1, column=5).value)) < 20:
                    new_ws_ul.cell(row=row_counter, column=j).value = round(
                        float(cur_ws.cell(row=row_start + 1, column=j).value - 5), 3)
                else:
                    new_ws_ul.cell(row=row_counter, column=j).value = "Fail"
            elif j == 3:
                if cur_ws.cell(row=row_end, column=5).value != "masked" and cur_ws.cell(row=row_end,
                                                                                        column=5).value != "mask" and \
                        abs(float(cur_ws.cell(row=row_end, column=5).value)) < 30:
                    new_ws_ul.cell(row=row_counter, column=j).value = round(
                        float(cur_ws.cell(row=row_end, column=j - 1).value), 3)
                elif cur_ws.cell(row=row_end - 1, column=5).value != "masked" and cur_ws.cell(row=row_end - 1,
                                                                                              column=5).value != "mask" and \
                        abs(float(cur_ws.cell(row=row_end - 1, column=5).value)) < 20:
                    new_ws_ul.cell(row=row_counter, column=j).value = round(
                        float(cur_ws.cell(row=row_end - 1, column=j - 1).value), 3)
                else:
                    new_ws_ul.cell(row=row_counter, column=j).value = "Fail"
        row_counter = row_counter + 1

    upper_and_lower_bond_list = []
    keys = []
    for j in range(1, 4, 1):
        keys.append(new_ws_ul.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 4, 1):
            row_data[keys[j - 1]] = new_ws_ul.cell(row=row_number, column=j).value
        upper_and_lower_bond_list.append(row_data)
    print(json.dumps(upper_and_lower_bond_list))


def wes_qc_data(folder_name):
    file_list = wes_extract_excels(folder_name)

    new_ws_qc = new_wb.create_sheet(title="QC Data Summary", index=3)
    new_ws_qc.cell(row=1, column=1).value = "RunNumber"
    new_ws_qc.cell(row=1, column=2).value = "QCIn1To10CSF"
    new_ws_qc.cell(row=1, column=3).value = "SpikedConcngPermL"
    new_ws_qc.cell(row=1, column=4).value = "ConcngPermL"
    new_ws_qc.cell(row=1, column=5).value = "PercentRE"

    row_counter = 2
    for file_name in file_list:
        # extract run number
        file_name_slice = file_name.split(".xlsx")
        this_run_number = int(file_name_slice[0].split()[-1])
        if file_name.split()[-3] == "masked":
            this_run_number = str(this_run_number) + " masked"
        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['Data Analysis']
        row_start = 16
        while cur_ws.cell(row=row_start, column=1).value is not None:
            for j in range(1, 6, 1):
                if j == 1:
                    new_ws_qc.cell(row=row_counter, column=j).value = this_run_number
                elif j == 2:
                    new_ws_qc.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 1).value
                else:
                    new_ws_qc.cell(row=row_counter, column=j).value = round(
                        float(cur_ws.cell(row=row_start, column=j - 1).value), 3)
            row_start = row_start + 1
            row_counter = row_counter + 1

    qc_data_list = []
    keys = []
    for j in range(1, 6, 1):
        keys.append(new_ws_qc.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 6, 1):
            row_data[keys[j - 1]] = new_ws_qc.cell(row=row_number, column=j).value
        qc_data_list.append(row_data)
    print(json.dumps(qc_data_list))


def wes_sample_analysis(folder_name):
    file_list = wes_extract_excels(folder_name)

    new_ws_sa = new_wb.create_sheet(title="Sample Analysis Data Summary", index=4)
    new_ws_sa.cell(row=1, column=1).value = "RunNumber"
    new_ws_sa.cell(row=1, column=2).value = "TimePoint"
    new_ws_sa.cell(row=1, column=3).value = "AnimalID"
    new_ws_sa.cell(row=1, column=4).value = "ConcngPermL"
    new_ws_sa.cell(row=1, column=5).value = "Dilution"
    new_ws_sa.cell(row=1, column=6).value = "AdjustedConcngPermL"
    new_ws_sa.cell(row=1, column=7).value = "Comment"

    row_counter = 2
    for file_name in file_list:
        # extract run number
        file_name_slice = file_name.split(".xlsx")
        this_run_number = int(file_name_slice[0].split()[-1])
        if file_name.split()[-3] == "masked":
            this_run_number = str(this_run_number) + " masked"
        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['Data Analysis']
        row_start = 21
        if this_run_number == 1:
            while cur_ws.cell(row=row_start, column=1).value is not None:
                for j in range(1, 8, 1):
                    if j == 1:
                        new_ws_sa.cell(row=row_counter, column=j).value = this_run_number
                    elif j == 4 or j == 6:
                        new_ws_sa.cell(row=row_counter, column=j).value = round(float(cur_ws.cell(row=row_start, column=j - 1).value), 3)
                    else:
                        new_ws_sa.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 1).value
                row_start = row_start + 1
                row_counter = row_counter + 1
        else:
            while cur_ws.cell(row=row_start, column=1).value is not None:
                comment = False
                time_info_list = []
                for j in range(1, 8, 1):
                    if j == 1:
                        new_ws_sa.cell(row=row_counter, column=j).value = this_run_number
                    elif j == 2:
                        time_info_list = str(cur_ws.cell(row=row_start, column=j - 1).value).split()
                        time_point = ""
                        for fragment in time_info_list:
                            if fragment == time_info_list[0]:
                                continue
                            else:
                                time_point = time_point + " " + fragment
                        new_ws_sa.cell(row=row_counter, column=j).value = time_point
                    elif j == 3:
                        new_ws_sa.cell(row=row_counter, column=j).value = int(time_info_list[0])
                    elif j == 4:
                        if cur_ws.cell(row=row_start, column=j - 1).value == "AQL" or \
                                cur_ws.cell(row=row_start, column=j - 1).value == "BLQ" or \
                                cur_ws.cell(row=row_start, column=j - 1).value == "BQL" or \
                                cur_ws.cell(row=row_start, column=j - 1).value == "ULQ":
                            cur_ws = cur_wb['Raw Data']
                            new_ws_sa.cell(row=row_counter, column=j).value = round(float(cur_ws.cell(row=row_start - 8,
                                                                                                      column=11).value),
                                                                                    3)
                            cur_ws = cur_wb['Data Analysis']
                            comment = True
                        else:
                            new_ws_sa.cell(row=row_counter, column=j).value = round(float(cur_ws.cell(row=row_start,
                                                                                                      column=j - 1).value),
                                                                                    3)
                    elif j == 6:
                        if cur_ws.cell(row=row_start, column=j - 1).value == "NA" or \
                                cur_ws.cell(row=row_start, column=j - 1).value == "BLQ":
                            new_ws_sa.cell(row=row_counter, column=j).value = round(
                                float(new_ws_sa.cell(row=row_counter, column=j - 1).value) *
                                float(new_ws_sa.cell(row=row_counter, column=j - 2).value), 3)
                        else:
                            new_ws_sa.cell(row=row_counter, column=j).value = round(float(cur_ws.cell(row=row_start,
                                                                                                      column=j - 1).value),
                                                                                    3)
                    elif j == 7:
                        if comment:
                            new_ws_sa.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start,
                                                                                          column=j - 4).value
                    else:
                        new_ws_sa.cell(row=row_counter, column=j).value = cur_ws.cell(row=row_start, column=j - 1).value

                row_start = row_start + 1
                row_counter = row_counter + 1

    sample_analysis_list = []
    keys = []
    for j in range(1, 8, 1):
        keys.append(new_ws_sa.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 8, 1):
            row_data[keys[j - 1]] = new_ws_sa.cell(row=row_number, column=j).value
        sample_analysis_list.append(row_data)
    print(json.dumps(sample_analysis_list))


def nab_data(folder_name):
    p = Path(folder_name)
    file_path_list = p.glob('Spark200*.xlsx')
    file_list = []
    for file_name in file_path_list:
        file_list.append(str(file_name))
    new_ws_nab = new_wb.create_sheet(title="NAB Summary", index=0)
    new_ws_nab.cell(row=1, column=1).value = "RunNumber"
    new_ws_nab.cell(row=1, column=2).value = "SampleNumber"
    new_ws_nab.cell(row=1, column=3).value = "SubjectID"
    new_ws_nab.cell(row=1, column=4).value = "VisitName"
    new_ws_nab.cell(row=1, column=5).value = "CollectionDate"
    new_ws_nab.cell(row=1, column=6).value = "MINAverage"
    new_ws_nab.cell(row=1, column=7).value = "MAXAverage"
    new_ws_nab.cell(row=1, column=8).value = "MAXEVAverage"
    new_ws_nab.cell(row=1, column=9).value = "FACT1To100Average"
    new_ws_nab.cell(row=1, column=10).value = "MINCVPercentage"
    new_ws_nab.cell(row=1, column=11).value = "MAXCVPercentage"
    new_ws_nab.cell(row=1, column=12).value = "MAXEVCVPercentage"
    new_ws_nab.cell(row=1, column=13).value = "FACT1To100CVPercentage"
    new_ws_nab.cell(row=1, column=14).value = "PercentageIMAXFACT1To100"
    new_ws_nab.cell(row=1, column=15).value = "EVInterferenceMAXMAXEV"
    new_ws_nab.cell(row=1, column=16).value = "EVEfficiencyFACT1To100FACT1To100EV"
    new_ws_nab.cell(row=1, column=17).value = "SN"
    new_ws_nab.cell(row=1, column=18).value = "S1To1IMAXPercentage"
    new_ws_nab.cell(row=1, column=19).value = "S1To2Point5IMAXPercentage"
    new_ws_nab.cell(row=1, column=20).value = "S1To5IMAXPercentage"
    new_ws_nab.cell(row=1, column=21).value = "S1To10IMAXPercentage"
    new_ws_nab.cell(row=1, column=22).value = "S1To100IMAXPercentage"
    new_ws_nab.cell(row=1, column=23).value = "S1To1000IMAXPercentage"
    new_ws_nab.cell(row=1, column=24).value = "S1To1IEVPercentage"
    new_ws_nab.cell(row=1, column=25).value = "S1To2Point5IEVPercentage"
    new_ws_nab.cell(row=1, column=26).value = "S1To5IEVPercentage"
    new_ws_nab.cell(row=1, column=27).value = "S1To10IEVPercentage"
    new_ws_nab.cell(row=1, column=28).value = "S1To100IEVPercentage"
    new_ws_nab.cell(row=1, column=29).value = "S1To1000IEVPercentage"
    new_ws_nab.cell(row=1, column=30).value = "NAbTiter"
    new_ws_nab.cell(row=1, column=31).value = "JustifyIfResultisInvalid"

    # create a counter for keep track on row number for the new xlsx file
    row_counter = 2

    for file_name in file_list:
        valid_spreadsheet = []
        sample_number = []
        # cur_wb = load_workbook(file_name, read_only=True, data_only=True)
        cur_wb = load_workbook(file_name, data_only=True)
        cur_ws = cur_wb['Sample information']
        row_start = 8
        spreadsheet_counter = 1
        while cur_ws.cell(row=row_start, column=2).value is not None:
            valid_spreadsheet.append('P' + str(spreadsheet_counter))
            sample_number.append('S' + str(spreadsheet_counter))
            row_start = row_start + 1
            spreadsheet_counter = spreadsheet_counter + 1

        for spreadsheet in valid_spreadsheet:
            cur_ws = cur_wb[spreadsheet]
            for j in range(1, 32, 1):
                if j == 1:
                    new_ws_nab.cell(row=row_counter, column=j).value = cur_ws['C3'].value
                elif j == 2:
                    new_ws_nab.cell(row=row_counter, column=j).value = sample_number[valid_spreadsheet.index(spreadsheet)]
                elif j == 3:
                    new_ws_nab.cell(row=row_counter, column=j).value = cur_ws['A6'].value
                elif j == 4:
                    new_ws_nab.cell(row=row_counter, column=j).value = cur_ws['C6'].value
                elif j == 5:
                    if isinstance(cur_ws['D6'].value, datetime.datetime):
                        extraction_date = str(cur_ws['D6'].value).split()
                        formal_date = datetime.datetime.strptime(extraction_date[0], '%Y-%m-%d').strftime('%d%b%Y')
                        new_ws_nab.cell(row=row_counter, column=j).value = formal_date
                    else:
                        new_ws_nab.cell(row=row_counter, column=j).value = str(cur_ws['D6'].value)
                elif 6 <= j <= 9:
                    new_ws_nab.cell(row=row_counter, column=j).value = \
                        round(float(cur_ws.cell(row=j + 26, column=3).value), 1)
                elif 10 <= j <= 13:
                    new_ws_nab.cell(row=row_counter, column=j).value = \
                        round(float(cur_ws.cell(row=j + 22, column=5).value), 1)
                elif 14 <= j <= 17:
                    new_ws_nab.cell(row=row_counter, column=j).value = \
                        round(float(cur_ws.cell(row=j + 18, column=9).value), 1)
                elif 18 <= j <= 23:
                    new_ws_nab.cell(row=row_counter, column=j).value = \
                        round(float(cur_ws.cell(row=j + 39, column=3).value))
                elif 24 <= j <= 29:
                    new_ws_nab.cell(row=row_counter, column=j).value = \
                        round(float(cur_ws.cell(row=j + 33, column=4).value))
                elif j == 31:
                    new_ws_nab.cell(row=row_counter, column=j).value = cur_ws['F57'].value
                else:
                    new_ws_nab.cell(row=row_counter, column=j).value = cur_ws['E57'].value
            row_counter = row_counter + 1

    nab_data_list = []
    keys = []
    for j in range(1, 32, 1):
        keys.append(new_ws_nab.cell(row=1, column=j).value)
    for row_number in range(1, row_counter):
        if row_number == 1:
            continue
        row_data = {}
        for j in range(1, 32, 1):
            row_data[keys[j - 1]] = new_ws_nab.cell(row=row_number, column=j).value
        nab_data_list.append(row_data)
    print(json.dumps(nab_data_list))


def main():
    if sys.argv[2] == 'qPCRqc':
        qpcr_qc_summary(sys.argv[1])
    elif sys.argv[2] == 'qPCRraw':
        qpcr_raw_data(sys.argv[1], sys.argv[2])
    elif sys.argv[2] == 'qPCRretest':
        qpcr_raw_data(sys.argv[1], sys.argv[2])
        qpcr_retest_fun()
    elif sys.argv[2] == 'qPCReachqc':
        qpcr_each_qc(sys.argv[1])
    elif sys.argv[2] == 'wesLinearRegressionDataSummary':
        wes_linear_regression(sys.argv[1])
    elif sys.argv[2] == 'wesStandardCurveDataSummary':
        wes_standard_curve(sys.argv[1])
    elif sys.argv[2] == 'wesUpperandLowerBondSummary':
        wes_upper_and_lower_bond(sys.argv[1])
    elif sys.argv[2] == 'wesQCDataSummary':
        wes_qc_data(sys.argv[1])
    elif sys.argv[2] == 'wesSampleAnalysisDataSummary':
        wes_sample_analysis(sys.argv[1])
    elif sys.argv[2] == 'nabDataSummary':
        nab_data(sys.argv[1])

if __name__ == '__main__':
    warnings.filterwarnings('ignore')
    main()
